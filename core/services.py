import csv
import csv
from calendar import monthrange
from datetime import timedelta
from decimal import Decimal
from io import TextIOWrapper

from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Avg, Count, Exists, Max, Min, OuterRef, Q, Sum
from django.utils import timezone
from django.utils.crypto import get_random_string
from openpyxl import load_workbook

from .models import AuditLog, Choice, Classroom, ContactMessage, Exam, Invoice, MpesaTransaction, Payment, PlatformAnnouncement, PlatformPricing, ProctorLog, Question, QuestionBankItem, QuestionSection, StudentAnswer, StudentNotification, StudentProfile, StudentReminder, StudentTodo, Submission, SubscriptionPlan, SupportIssue, TutorProfile, TutorSubscription, User


def build_unique_username(email):
    """Create a stable unique username from an email address."""
    UserModel = get_user_model()
    base_username = email.split("@", 1)[0].lower()[:140] or "user"
    username = base_username
    counter = 1
    while UserModel.objects.filter(username=username).exists():
        counter += 1
        suffix = f"-{counter}"
        username = f"{base_username[:150 - len(suffix)]}{suffix}"
    return username


def create_tutor_account(*, fullname, email, password, phone_number="", institution_name=""):
    """Create a self-registered tutor and matching tutor profile."""
    user = User.objects.create_user(
        username=build_unique_username(email),
        email=email,
        password=password,
        fullname=fullname,
        phone_number=phone_number,
        role=User.ROLE_TUTOR,
    )
    TutorProfile.objects.create(user=user, institution_name=institution_name)
    return user


def create_student_account(*, tutor, fullname, email, school_name, registration_number):
    """Provision a student account owned by a tutor.

    The generated password is returned once so the tutor can share it. Students
    may also log in using their registration number during the early workflow.
    """
    temporary_password = get_random_string(10)
    user = User.objects.create_user(
        username=build_unique_username(email),
        email=email,
        password=temporary_password,
        fullname=fullname,
        role=User.ROLE_STUDENT,
        school_name=school_name,
        registration_id=registration_number,
    )
    StudentProfile.objects.create(
        user=user,
        tutor=tutor,
        school_name=school_name,
        registration_number=registration_number,
    )
    return user, temporary_password


def generate_classroom_code():
    """Create a short join/reference code for a classroom."""
    code = get_random_string(8).upper()
    while Classroom.objects.filter(room_id=code).exists():
        code = get_random_string(8).upper()
    return code


def create_classroom(*, tutor, name):
    """Create a tutor-owned classroom with a non-public legacy password value."""
    return Classroom.objects.create(
        tutor=tutor,
        name=name,
        room_id=generate_classroom_code(),
        password=get_random_string(12),
    )


def update_classroom(*, tutor, classroom, name):
    """Update a tutor-owned classroom."""
    if classroom.tutor_id != tutor.id:
        raise PermissionError("Classroom is not owned by this tutor.")
    classroom.name = name
    classroom.save(update_fields=["name", "updated_at"])
    return classroom


def archive_classroom(*, tutor, classroom):
    """Archive a tutor-owned classroom without deleting history."""
    if classroom.tutor_id != tutor.id:
        raise PermissionError("Classroom is not owned by this tutor.")
    classroom.is_archived = True
    classroom.save(update_fields=["is_archived", "updated_at"])
    return classroom


def assign_student_to_classroom(*, tutor, classroom, student):
    """Attach a tutor-owned student to a tutor-owned classroom."""
    if classroom.tutor_id != tutor.id:
        raise PermissionError("Classroom is not owned by this tutor.")
    if not StudentProfile.objects.filter(tutor=tutor, user=student).exists():
        raise PermissionError("Student is not owned by this tutor.")
    classroom.students.add(student)
    return classroom


def assign_existing_student_by_registration(*, tutor, classroom, registration_number):
    """Attach an existing tutor-owned student by registration number.

    Returns the student user and whether a new classroom membership was created.
    """
    if classroom.tutor_id != tutor.id:
        raise PermissionError("Classroom is not owned by this tutor.")
    profile = StudentProfile.objects.select_related("user").get(
        tutor=tutor,
        registration_number=registration_number,
    )
    already_assigned = classroom.students.filter(id=profile.user_id).exists()
    if not already_assigned:
        classroom.students.add(profile.user)
    return profile.user, not already_assigned


def remove_student_from_classroom(*, tutor, classroom, student):
    """Remove a tutor-owned student from a classroom without deleting the account."""
    if classroom.tutor_id != tutor.id:
        raise PermissionError("Classroom is not owned by this tutor.")
    if not StudentProfile.objects.filter(tutor=tutor, user=student).exists():
        raise PermissionError("Student is not owned by this tutor.")
    classroom.students.remove(student)
    return classroom


def reset_student_credentials(*, tutor, student):
    """Reset a tutor-owned student's temporary password.

    Ownership is checked here so views and future APIs use the same guardrail.
    """
    try:
        profile = StudentProfile.objects.select_related("user").get(user=student, tutor=tutor)
    except StudentProfile.DoesNotExist as exc:
        raise PermissionError("Student is not owned by this tutor.") from exc
    temporary_password = get_random_string(10)
    profile.user.set_password(temporary_password)
    profile.user.save(update_fields=["password"])
    profile.must_change_password = True
    profile.save(update_fields=["must_change_password"])
    return temporary_password


def import_students_from_file(*, tutor, upload, classroom=None):
    """Provision students from CSV/XLSX rows and optionally assign a classroom.

    Expected columns are `fullname`, `email`, `school_name`, and
    `registration_number`. The result carries created credentials so tutors can
    share them immediately after import.
    """
    rows = _read_student_rows(upload)
    result = {"created": [], "skipped": [], "errors": []}

    for row_number, row in enumerate(rows, start=2):
        fullname = (row.get("fullname") or "").strip()
        email = (row.get("email") or "").strip().lower()
        school_name = (row.get("school_name") or "").strip()
        registration_number = (row.get("registration_number") or "").strip()

        if not all([fullname, email, school_name, registration_number]):
            result["errors"].append(f"Row {row_number}: missing required data.")
            continue
        if User.objects.filter(email=email).exists():
            result["skipped"].append(f"Row {row_number}: {email} already exists.")
            continue
        if StudentProfile.objects.filter(tutor=tutor, registration_number=registration_number).exists():
            result["skipped"].append(f"Row {row_number}: registration {registration_number} already exists.")
            continue

        student, temporary_password = create_student_account(
            tutor=tutor,
            fullname=fullname,
            email=email,
            school_name=school_name,
            registration_number=registration_number,
        )
        if classroom:
            assign_student_to_classroom(tutor=tutor, classroom=classroom, student=student)
        result["created"].append(
            {
                "name": student.fullname,
                "email": student.email,
                "registration_number": registration_number,
                "temporary_password": temporary_password,
            }
        )

    return result


def _read_student_rows(upload):
    """Read normalized dict rows from a CSV or XLSX upload."""
    name = upload.name.lower()
    upload.seek(0)
    if name.endswith(".csv"):
        wrapper = TextIOWrapper(upload.file, encoding="utf-8-sig")
        return list(csv.DictReader(wrapper))

    workbook = load_workbook(upload, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip().lower() if value is not None else "" for value in rows[0]]
    normalized = []
    for values in rows[1:]:
        normalized.append(
            {
                headers[index]: "" if value is None else str(value)
                for index, value in enumerate(values)
                if index < len(headers)
            }
        )
    return normalized


def create_assessment(*, tutor, data):
    """Create a draft assessment owned through the tutor's classroom."""
    classroom = Classroom.objects.get(id=data["classroom_id"], tutor=tutor, is_archived=False)
    return Exam.objects.create(classroom=classroom, **_assessment_kwargs(data))


def update_assessment(*, tutor, exam, data):
    """Update a tutor-owned draft/published assessment."""
    _ensure_assessment_owner(tutor=tutor, exam=exam)
    classroom = Classroom.objects.get(id=data["classroom_id"], tutor=tutor, is_archived=False)
    for field, value in _assessment_kwargs(data).items():
        setattr(exam, field, value)
    exam.classroom = classroom
    exam.save()
    return exam


def publish_assessment(*, tutor, exam):
    """Publish an assessment after validating it has at least one question."""
    _ensure_assessment_owner(tutor=tutor, exam=exam)
    if not exam.questions.exists():
        raise ValueError("Add at least one question before publishing.")
    exam.status = Exam.STATUS_PUBLISHED
    exam.save(update_fields=["status", "updated_at"])
    return exam


def get_platform_pricing():
    """Return active platform pricing, creating the default KES 5 rate if needed."""
    pricing = PlatformPricing.objects.filter(is_active=True).order_by("-updated_at").first()
    if pricing:
        return pricing
    return PlatformPricing.objects.create(currency="KES", pay_per_student_rate=Decimal("5.00"))


def get_or_create_assessment_invoice(*, tutor, exam):
    """Create or return the pending per-student invoice for an assessment."""
    _ensure_assessment_owner(tutor=tutor, exam=exam)
    paid_invoice = Invoice.objects.filter(
        tutor=tutor,
        assessment=exam,
        invoice_type=Invoice.TYPE_ASSESSMENT,
        status=Invoice.STATUS_PAID,
    ).first()
    if paid_invoice:
        return paid_invoice

    invoice = Invoice.objects.filter(
        tutor=tutor,
        assessment=exam,
        invoice_type=Invoice.TYPE_ASSESSMENT,
        status__in=[Invoice.STATUS_DRAFT, Invoice.STATUS_PENDING],
    ).first()
    pricing = get_platform_pricing()
    quantity = exam.classroom.students.count()
    subtotal = pricing.pay_per_student_rate * Decimal(quantity)
    if invoice:
        invoice.currency = pricing.currency
        invoice.unit_amount = pricing.pay_per_student_rate
        invoice.quantity = quantity
        invoice.subtotal = subtotal
        invoice.discount_amount = Decimal("0.00")
        invoice.total_amount = subtotal
        invoice.status = Invoice.STATUS_PENDING
        invoice.save(
            update_fields=[
                "currency",
                "unit_amount",
                "quantity",
                "subtotal",
                "discount_amount",
                "total_amount",
                "status",
            ]
        )
        return invoice

    return Invoice.objects.create(
        tutor=tutor,
        assessment=exam,
        invoice_type=Invoice.TYPE_ASSESSMENT,
        currency=pricing.currency,
        unit_amount=pricing.pay_per_student_rate,
        quantity=quantity,
        subtotal=subtotal,
        total_amount=subtotal,
        notes="Per-student assessment access invoice.",
    )


def create_subscription_invoice(*, tutor, plan):
    """Create a pending invoice for a tutor subscription plan.

    A tutor cannot buy the same active plan again. During an active
    subscription they may only move to a higher-priced plan. When they choose
    an upgrade, unused value from the current plan is carried as a credit on
    the new invoice.
    """
    active_subscription = get_active_subscription(tutor=tutor)
    if active_subscription and active_subscription.plan_id == plan.id:
        raise ValueError("You already have this subscription plan active.")
    if active_subscription and not _is_subscription_upgrade(current_plan=active_subscription.plan, new_plan=plan):
        raise ValueError("You can only upgrade while your current subscription is active. You can choose a lower plan after the current plan ends.")

    subtotal = plan.price
    discount_amount = (subtotal * plan.discount_percent / Decimal("100")).quantize(Decimal("0.01"))
    upgrade_credit = estimate_subscription_upgrade_credit(tutor=tutor, new_plan=plan)
    discount_amount = min(subtotal, discount_amount + upgrade_credit)
    total_amount = subtotal - discount_amount
    matching_invoice = _matching_pending_subscription_invoice(
        tutor=tutor,
        plan=plan,
        currency=plan.currency,
        subtotal=subtotal,
        discount_amount=discount_amount,
        total_amount=total_amount,
    )
    if matching_invoice:
        return matching_invoice

    return Invoice.objects.create(
        tutor=tutor,
        subscription_plan=plan,
        invoice_type=Invoice.TYPE_SUBSCRIPTION,
        currency=plan.currency,
        unit_amount=plan.price,
        quantity=1,
        subtotal=subtotal,
        discount_amount=discount_amount,
        total_amount=total_amount,
        notes=_subscription_invoice_notes(plan=plan, upgrade_credit=upgrade_credit),
    )


def estimate_subscription_upgrade_credit(*, tutor, new_plan):
    """Estimate unused current-plan value to deduct from a different plan."""
    active_subscription = get_active_subscription(tutor=tutor)
    if not active_subscription or active_subscription.plan_id == new_plan.id:
        return Decimal("0.00")
    if not _is_subscription_upgrade(current_plan=active_subscription.plan, new_plan=new_plan):
        return Decimal("0.00")
    if not active_subscription.starts_at or not active_subscription.ends_at:
        return Decimal("0.00")
    now = timezone.now()
    total_seconds = Decimal(str((active_subscription.ends_at - active_subscription.starts_at).total_seconds()))
    remaining_seconds = Decimal(str(max(0, (active_subscription.ends_at - now).total_seconds())))
    if total_seconds <= 0 or remaining_seconds <= 0:
        return Decimal("0.00")
    credit = active_subscription.plan.price * (remaining_seconds / total_seconds)
    return min(new_plan.price, credit.quantize(Decimal("0.01")))


def _is_subscription_upgrade(*, current_plan, new_plan):
    """Return True when the new plan is higher value than the current plan."""
    return new_plan.price > current_plan.price


def _matching_pending_subscription_invoice(*, tutor, plan, currency, subtotal, discount_amount, total_amount):
    """Reuse an unpaid matching invoice instead of duplicating the same plan/price."""
    return Invoice.objects.filter(
        tutor=tutor,
        subscription_plan=plan,
        invoice_type=Invoice.TYPE_SUBSCRIPTION,
        status__in=[Invoice.STATUS_DRAFT, Invoice.STATUS_PENDING],
        currency=currency,
        unit_amount=plan.price,
        quantity=1,
        subtotal=subtotal,
        discount_amount=discount_amount,
        total_amount=total_amount,
    ).order_by("-created_at").first()


def _subscription_invoice_notes(*, plan, upgrade_credit):
    """Describe subscription billing adjustments for the invoice record."""
    notes = f"{plan.name} subscription for {plan.duration_months} month(s)."
    if upgrade_credit:
        notes += f" Includes KES {upgrade_credit} unused subscription credit."
    return notes


def get_active_subscription(*, tutor):
    """Return the current active tutor subscription, if one exists."""
    now = timezone.now()
    return TutorSubscription.objects.filter(
        tutor=tutor,
        status=TutorSubscription.STATUS_ACTIVE,
        starts_at__lte=now,
        ends_at__gte=now,
    ).select_related("plan").order_by("-ends_at").first()


def assessment_can_publish_without_invoice(*, tutor, exam):
    """Return True when subscription or zero-student invoice allows publishing."""
    _ensure_assessment_owner(tutor=tutor, exam=exam)
    if get_active_subscription(tutor=tutor):
        return True
    invoice = get_or_create_assessment_invoice(tutor=tutor, exam=exam)
    return invoice.status == Invoice.STATUS_PAID or invoice.total_amount == Decimal("0.00")


@transaction.atomic
def mark_invoice_paid(*, invoice, method=Payment.METHOD_MANUAL, provider_reference=""):
    """Record a successful payment and activate linked subscriptions."""
    invoice = Invoice.objects.select_for_update().get(id=invoice.id)
    if invoice.status == Invoice.STATUS_PAID:
        return invoice.payments.filter(status=Payment.STATUS_SUCCESSFUL).order_by("-confirmed_at").first()

    now = timezone.now()
    payment = Payment.objects.create(
        invoice=invoice,
        tutor=invoice.tutor,
        amount=invoice.total_amount,
        currency=invoice.currency,
        method=method,
        status=Payment.STATUS_SUCCESSFUL,
        provider_reference=provider_reference,
        confirmed_at=now,
    )
    invoice.status = Invoice.STATUS_PAID
    invoice.paid_at = now
    invoice.save(update_fields=["status", "paid_at"])
    if invoice.invoice_type == Invoice.TYPE_SUBSCRIPTION and invoice.subscription_plan:
        activate_subscription_from_invoice(invoice=invoice)
    return payment


def activate_subscription_from_invoice(*, invoice):
    """Start or extend a tutor subscription after its invoice is paid."""
    if invoice.status != Invoice.STATUS_PAID or not invoice.subscription_plan:
        return None
    now = timezone.now()
    current = get_active_subscription(tutor=invoice.tutor)
    is_upgrade = bool(current and current.plan_id != invoice.subscription_plan_id)
    starts_at = now
    if is_upgrade:
        current.status = TutorSubscription.STATUS_CANCELLED
        current.ends_at = now
        current.save(update_fields=["status", "ends_at"])
    ends_at = _add_months(starts_at, invoice.subscription_plan.duration_months)
    return TutorSubscription.objects.update_or_create(
        invoice=invoice,
        defaults={
            "tutor": invoice.tutor,
            "plan": invoice.subscription_plan,
            "status": TutorSubscription.STATUS_ACTIVE,
            "starts_at": starts_at,
            "ends_at": ends_at,
        },
    )[0]


def create_mpesa_payment_attempt(*, invoice, phone_number=""):
    """Create a local pending M-Pesa payment before calling STK Push."""
    return Payment.objects.create(
        invoice=invoice,
        tutor=invoice.tutor,
        amount=invoice.total_amount,
        currency=invoice.currency,
        method=Payment.METHOD_MPESA,
        status=Payment.STATUS_PENDING,
        provider_reference=phone_number,
    )


@transaction.atomic
def handle_mpesa_stk_callback(payload):
    """Reconcile a Daraja STK callback idempotently."""
    callback = payload.get("Body", {}).get("stkCallback", {})
    checkout_request_id = callback.get("CheckoutRequestID", "")
    merchant_request_id = callback.get("MerchantRequestID", "")
    result_code = callback.get("ResultCode")
    result_description = callback.get("ResultDesc", "")
    if not checkout_request_id:
        raise ValueError("M-Pesa STK callback is missing CheckoutRequestID.")
    existing = MpesaTransaction.objects.filter(checkout_request_id=checkout_request_id).first()
    if existing:
        return existing

    metadata = _mpesa_metadata(callback)
    payment = Payment.objects.filter(checkout_request_id=checkout_request_id).select_related("invoice").first()
    transaction_record = MpesaTransaction.objects.create(
        payment=payment,
        merchant_request_id=merchant_request_id,
        checkout_request_id=checkout_request_id,
        receipt_number=metadata.get("MpesaReceiptNumber", ""),
        phone_number=str(metadata.get("PhoneNumber", "")),
        amount=Decimal(str(metadata.get("Amount") or "0")),
        result_code=result_code,
        result_description=result_description,
        raw_payload=payload,
    )
    if payment and result_code == 0:
        payment.status = Payment.STATUS_SUCCESSFUL
        payment.provider_reference = transaction_record.receipt_number
        payment.merchant_request_id = merchant_request_id
        payment.confirmed_at = timezone.now()
        payment.save(
            update_fields=[
                "status",
                "provider_reference",
                "merchant_request_id",
                "confirmed_at",
            ]
        )
        invoice = payment.invoice
        invoice.status = Invoice.STATUS_PAID
        invoice.paid_at = payment.confirmed_at
        invoice.save(update_fields=["status", "paid_at"])
        if invoice.invoice_type == Invoice.TYPE_SUBSCRIPTION:
            activate_subscription_from_invoice(invoice=invoice)
    elif payment:
        payment.status = Payment.STATUS_FAILED
        payment.merchant_request_id = merchant_request_id
        payment.save(update_fields=["status", "merchant_request_id"])
    return transaction_record


def _mpesa_metadata(callback):
    """Flatten Daraja callback metadata items by their `Name` values."""
    items = callback.get("CallbackMetadata", {}).get("Item", [])
    return {item.get("Name"): item.get("Value") for item in items}


def _add_months(value, months):
    """Add calendar months while preserving the closest valid day."""
    month_index = value.month - 1 + months
    year = value.year + month_index // 12
    month = month_index % 12 + 1
    day = min(value.day, monthrange(year, month)[1])
    return value.replace(year=year, month=month, day=day)


def get_student_assessment_overview(*, student):
    """Return published assessments assigned to a student, grouped for dashboard display."""
    now = timezone.now()
    completed_submissions = Submission.objects.filter(
        student=student,
        exam=OuterRef("pk"),
        completed=True,
    )
    assessments = (
        Exam.objects.filter(status=Exam.STATUS_PUBLISHED, classroom__students=student)
        .select_related("classroom", "classroom__tutor")
        .annotate(has_completed_submission=Exists(completed_submissions))
        .order_by("end_time", "start_time", "title")
        .distinct()
    )
    overview = {
        "pending": [],
        "active": [],
        "overdue": [],
        "completed": [],
        "calendar": [],
    }
    for assessment in assessments:
        status = _student_assessment_status(assessment=assessment, now=now)
        if assessment.has_completed_submission:
            status = "completed"
        overview[status].append(assessment)
        overview["calendar"].append(
            {
                "assessment": assessment,
                "status": status,
                "date": assessment.start_time or assessment.end_time,
            }
        )
    overview["calendar"] = [
        item for item in overview["calendar"] if item["date"]
    ][:10]
    return overview


def create_student_todo(*, student, data):
    """Create a student-owned todo and optional reminder."""
    assessment = None
    if data.get("assessment_id"):
        assessment = Exam.objects.get(
            id=data["assessment_id"],
            status=Exam.STATUS_PUBLISHED,
            classroom__students=student,
        )
    todo = StudentTodo.objects.create(
        student=student,
        title=data["title"],
        notes=data.get("notes", ""),
        due_at=data.get("due_at"),
        assessment=assessment,
    )
    if todo.due_at:
        StudentReminder.objects.create(
            student=student,
            todo=todo,
            assessment=assessment,
            title=f"Todo due: {todo.title}",
            remind_at=todo.due_at,
            kind=StudentReminder.KIND_TODO,
        )
    return todo


def complete_student_todo(*, student, todo):
    """Mark a todo complete only when it belongs to the student."""
    if todo.student_id != student.id:
        raise PermissionError("Todo is not owned by this student.")
    todo.mark_complete()
    return todo


def mark_student_notification_read(*, student, notification):
    """Mark one student notification as read with ownership protection."""
    if notification.student_id != student.id:
        raise PermissionError("Notification is not owned by this student.")
    notification.mark_read()
    return notification


def sync_student_assessment_notifications(*, student):
    """Create one in-app notification for each newly visible assessment."""
    assessments = Exam.objects.filter(status=Exam.STATUS_PUBLISHED, classroom__students=student)
    for assessment in assessments:
        StudentNotification.objects.get_or_create(
            student=student,
            assessment=assessment,
            notification_type=StudentNotification.TYPE_ASSESSMENT,
            defaults={
                "title": f"New {assessment.get_assessment_type_display()}: {assessment.title}",
                "message": f"{assessment.classroom.name} has a published assessment ready for you.",
            },
        )


def get_student_dashboard_context(*, student):
    """Build the complete student dashboard context."""
    sync_student_assessment_notifications(student=student)
    overview = get_student_assessment_overview(student=student)
    visible_assessments = (
        Exam.objects.filter(status=Exam.STATUS_PUBLISHED, classroom__students=student)
        .order_by("title")
        .distinct()
    )
    return {
        "assessment_overview": overview,
        "assessment_choices": visible_assessments,
        "todos": StudentTodo.objects.filter(student=student).select_related("assessment")[:10],
        "reminders": StudentReminder.objects.filter(student=student, is_sent=False).select_related("assessment", "todo")[:10],
        "notifications": StudentNotification.objects.filter(student=student).select_related("assessment")[:10],
        "recent_results": Submission.objects.filter(
            student=student,
            completed=True,
            exam__show_results_immediately=True,
        ).select_related("exam").order_by("-submitted_at")[:10],
        "unread_notification_count": StudentNotification.objects.filter(student=student, is_read=False).count(),
        "summary": {
            "pending_count": len(overview["pending"]),
            "active_count": len(overview["active"]),
            "overdue_count": len(overview["overdue"]),
            "completed_count": len(overview["completed"]),
            "todo_count": StudentTodo.objects.filter(student=student, is_completed=False).count(),
        },
    }


def _student_assessment_status(*, assessment, now):
    """Classify an assigned assessment for the student dashboard."""
    if assessment.start_time and assessment.start_time > now:
        return "pending"
    if assessment.end_time and assessment.end_time < now and not assessment.allow_late_submission:
        return "overdue"
    return "active"


def get_student_assessment_for_taking(*, student, exam_id):
    """Return an assessment only when it is assigned and visible to the student."""
    return Exam.objects.select_related("classroom").prefetch_related(
        "sections",
        "questions__choices",
    ).get(
        id=exam_id,
        status=Exam.STATUS_PUBLISHED,
        classroom__students=student,
    )


def get_or_start_submission(*, student, exam):
    """Create or resume the student's current attempt for an assessment."""
    _ensure_student_can_attempt(student=student, exam=exam)
    open_submission = Submission.objects.filter(student=student, exam=exam, completed=False).order_by("-started_at").first()
    if open_submission:
        return open_submission

    completed_attempts = Submission.objects.filter(student=student, exam=exam, completed=True).count()
    if completed_attempts >= exam.attempts_allowed:
        raise ValueError("You have used all allowed attempts for this assessment.")
    attempt_number = completed_attempts + 1
    started_at = timezone.now()
    expires_at = _submission_expiry(started_at=started_at, exam=exam)
    return Submission.objects.create(
        student=student,
        exam=exam,
        attempt_number=attempt_number,
        expires_at=expires_at,
        last_saved_at=started_at,
    )


def get_submission_time_remaining(*, submission):
    """Return remaining seconds before an attempt expires."""
    if not submission.expires_at:
        return None
    remaining = int((submission.expires_at - timezone.now()).total_seconds())
    return max(0, remaining)


def save_submission_answers(*, submission, answer_data):
    """Persist posted answer values without completing the attempt."""
    if submission.completed:
        raise ValueError("This attempt has already been submitted.")
    if _submission_is_expired(submission):
        raise ValueError("Time is up for this attempt.")
    for question in submission.exam.questions.all():
        value = answer_data.get(str(question.id), "")
        if isinstance(value, list):
            value = ",".join(sorted([str(item).strip().upper() for item in value if str(item).strip()]))
        else:
            value = str(value).strip()
        StudentAnswer.objects.update_or_create(
            submission=submission,
            question=question,
            defaults={"selected_choices": value},
        )
    submission.last_saved_at = timezone.now()
    submission.save(update_fields=["last_saved_at"])
    return submission


@transaction.atomic
def submit_assessment_attempt(*, student, submission, answer_data=None, auto_submitted=False):
    """Save final answers, grade objective questions, and complete an attempt."""
    submission = Submission.objects.select_for_update().select_related("exam").get(id=submission.id, student=student)
    if submission.completed:
        return submission
    if answer_data is not None and not _submission_is_expired(submission):
        save_submission_answers(submission=submission, answer_data=answer_data)

    score = Decimal("0.00")
    for question in submission.exam.questions.prefetch_related("choices"):
        answer, _ = StudentAnswer.objects.get_or_create(submission=submission, question=question)
        awarded_marks, is_correct = _grade_question_answer(question=question, raw_answer=answer.selected_choices)
        answer.awarded_marks = awarded_marks
        answer.is_correct = is_correct
        answer.save(update_fields=["awarded_marks", "is_correct", "saved_at"])
        score += awarded_marks

    submission.score = float(score)
    submission.completed = True
    submission.submitted_at = timezone.now()
    submission.last_saved_at = submission.submitted_at
    submission.save(update_fields=["score", "completed", "submitted_at", "last_saved_at"])
    if auto_submitted:
        StudentNotification.objects.create(
            student=student,
            assessment=submission.exam,
            title=f"Auto-submitted: {submission.exam.title}",
            message="Your attempt was submitted because the timer ended.",
            notification_type=StudentNotification.TYPE_SYSTEM,
        )
    return submission


def build_attempt_context(*, submission):
    """Prepare question and answer data for the exam-taking template."""
    questions = list(submission.exam.questions.select_related("section").prefetch_related("choices").order_by("section__order", "order", "id"))
    if submission.exam.randomize_questions and not submission.answers.exists():
        questions = sorted(questions, key=lambda question: question.id)
    answers = {
        answer.question_id: answer.selected_choices
        for answer in submission.answers.all()
    }
    for question in questions:
        question.saved_answer = answers.get(question.id, "")
        question.saved_answer_labels = _normalize_answer_labels(question.saved_answer).split(",") if question.saved_answer else []
    return {
        "submission": submission,
        "assessment": submission.exam,
        "questions": questions,
        "answers": answers,
        "time_remaining_seconds": get_submission_time_remaining(submission=submission),
        "proctoring": get_proctoring_context(submission=submission),
    }


def get_proctoring_context(*, submission):
    """Return serialized proctoring rules for the attempt page."""
    exam = submission.exam
    violation_count = ProctorLog.objects.filter(submission=submission).count()
    return {
        "enabled": exam.proctoring_enabled,
        "disable_copy_paste": exam.disable_copy_paste,
        "disable_right_click": exam.disable_right_click,
        "disable_text_selection": exam.disable_text_selection,
        "detect_tab_switch": exam.detect_tab_switch,
        "detect_window_blur": exam.detect_window_blur,
        "require_fullscreen": exam.require_fullscreen,
        "detect_fullscreen_exit": exam.detect_fullscreen_exit,
        "detect_refresh": exam.detect_refresh,
        "max_violation_warnings": exam.max_violation_warnings,
        "auto_submit_on_violation": exam.auto_submit_on_violation,
        "auto_disqualify_on_violation": exam.auto_disqualify_on_violation,
        "violation_count": violation_count,
    }


@transaction.atomic
def record_proctor_violation(*, student, submission, violation_type, details=""):
    """Log a proctoring violation and apply configured threshold actions."""
    submission = Submission.objects.select_for_update().select_related("exam").get(id=submission.id, student=student)
    exam = submission.exam
    if not exam.proctoring_enabled or submission.completed:
        return None, {"ignored": True}
    violation_count = ProctorLog.objects.filter(submission=submission).count() + 1
    should_act = violation_count >= exam.max_violation_warnings
    log = ProctorLog.objects.create(
        student=student,
        exam=exam,
        submission=submission,
        violation_type=violation_type,
        details=details,
        violation_count=violation_count,
        triggered_disqualification=should_act and exam.auto_disqualify_on_violation,
    )
    action = {
        "ignored": False,
        "violation_count": violation_count,
        "max_violation_warnings": exam.max_violation_warnings,
        "disqualified": False,
        "submitted": False,
    }
    if should_act and exam.auto_disqualify_on_violation:
        submission.is_disqualified = True
        submission.disqualified_at = timezone.now()
        submission.disqualification_reason = f"Exceeded violation threshold with {violation_type}."
        submission.completed = True
        submission.submitted_at = submission.disqualified_at
        submission.last_saved_at = submission.disqualified_at
        submission.save(
            update_fields=[
                "is_disqualified",
                "disqualified_at",
                "disqualification_reason",
                "completed",
                "submitted_at",
                "last_saved_at",
            ]
        )
        action["disqualified"] = True
    elif should_act and exam.auto_submit_on_violation:
        submit_assessment_attempt(student=student, submission=submission, answer_data=None, auto_submitted=True)
        action["submitted"] = True
    return log, action


def _ensure_student_can_attempt(*, student, exam):
    """Raise when an assessment cannot currently be attempted by a student."""
    now = timezone.now()
    if exam.status != Exam.STATUS_PUBLISHED:
        raise ValueError("This assessment is not available.")
    if not exam.classroom.students.filter(id=student.id).exists():
        raise PermissionError("Assessment is not assigned to this student.")
    if exam.start_time and exam.start_time > now:
        raise ValueError("This assessment has not started yet.")
    if exam.end_time and exam.end_time < now and not exam.allow_late_submission:
        raise ValueError("This assessment is closed.")


def _submission_expiry(*, started_at, exam):
    """Calculate the strict attempt expiry from duration and assessment end time."""
    duration_expiry = started_at + timedelta(minutes=exam.duration_minutes)
    if exam.end_time and not exam.allow_late_submission:
        return min(duration_expiry, exam.end_time)
    return duration_expiry


def _submission_is_expired(submission):
    """Return True if the attempt timer has elapsed."""
    return bool(submission.expires_at and submission.expires_at <= timezone.now())


def _grade_question_answer(*, question, raw_answer):
    """Grade objective questions and leave written/manual questions at zero."""
    objective_types = {
        Question.TYPE_SINGLE_CHOICE,
        Question.TYPE_MULTIPLE_CHOICE,
        Question.TYPE_TRUE_FALSE,
    }
    if question.question_type not in objective_types or not question.correct_labels:
        return Decimal("0.00"), False
    expected = _normalize_answer_labels(question.correct_labels)
    actual = _normalize_answer_labels(raw_answer)
    is_correct = expected == actual
    return (question.marks if is_correct else Decimal("0.00")), is_correct


def _normalize_answer_labels(value):
    """Normalize comma-separated choice labels for comparison."""
    return ",".join(sorted({part.strip().upper() for part in str(value).split(",") if part.strip()}))


def _assessment_kwargs(data):
    """Map validated form data into Exam fields."""
    return {
        "title": data["title"],
        "assessment_type": data["assessment_type"],
        "instructions": data.get("instructions", ""),
        "start_time": data.get("start_time"),
        "end_time": data.get("end_time"),
        "duration_minutes": data["duration_minutes"],
        "total_marks": data["total_marks"],
        "pass_mark": data["pass_mark"],
        "attempts_allowed": data["attempts_allowed"],
        "back_btn_enabled": data.get("back_btn_enabled", False),
        "allow_late_submission": data.get("allow_late_submission", False),
        "show_results_immediately": data.get("show_results_immediately", False),
        "show_answers": data.get("show_answers", False),
        "randomize_questions": data.get("randomize_questions", False),
        "randomize_choices": data.get("randomize_choices", False),
        "proctoring_enabled": data.get("proctoring_enabled", False),
        "disable_copy_paste": data.get("disable_copy_paste", False),
        "disable_right_click": data.get("disable_right_click", False),
        "disable_text_selection": data.get("disable_text_selection", False),
        "detect_tab_switch": data.get("detect_tab_switch", False),
        "detect_window_blur": data.get("detect_window_blur", False),
        "require_fullscreen": data.get("require_fullscreen", False),
        "detect_fullscreen_exit": data.get("detect_fullscreen_exit", False),
        "detect_refresh": data.get("detect_refresh", False),
        "max_violation_warnings": data.get("max_violation_warnings") or 3,
        "auto_submit_on_violation": data.get("auto_submit_on_violation", False),
        "auto_disqualify_on_violation": data.get("auto_disqualify_on_violation", False),
    }


def create_question_section(*, tutor, exam, data):
    """Create an ordered section in a tutor-owned assessment."""
    _ensure_assessment_owner(tutor=tutor, exam=exam)
    return QuestionSection.objects.create(
        exam=exam,
        title=data["title"],
        instructions=data.get("instructions", ""),
        order=data.get("order") or 0,
    )


def create_question_from_form(*, tutor, exam, data):
    """Create a question with choices and optionally save it to the bank."""
    _ensure_assessment_owner(tutor=tutor, exam=exam)
    section = None
    if data.get("section_id"):
        section = QuestionSection.objects.get(id=data["section_id"], exam=exam)
    question = Question.objects.create(
        exam=exam,
        section=section,
        question_type=data["question_type"],
        text=data["text"],
        explanation=data.get("explanation", ""),
        media_url=data.get("media_url", ""),
        marks=data["marks"],
        correct_labels=data.get("correct_labels", ""),
        order=data.get("order") or _next_question_order(exam),
        reusable_in_bank=data.get("reusable_in_bank", False),
    )
    choices = _choices_from_data(data)
    _create_choices(question=question, choices=choices)
    if question.reusable_in_bank:
        save_question_to_bank(tutor=tutor, question=question)
    _sync_assessment_total_marks(exam)
    return question


def save_question_to_bank(*, tutor, question):
    """Persist a reusable copy of a question for future assessments."""
    _ensure_assessment_owner(tutor=tutor, exam=question.exam)
    return QuestionBankItem.objects.create(
        tutor=tutor,
        title=question.text[:120],
        question_type=question.question_type,
        text=question.text,
        explanation=question.explanation,
        media_url=question.media_url,
        marks=question.marks,
        correct_labels=question.correct_labels,
        choices_json=list(question.choices.values("label", "text")),
    )


def add_bank_question_to_assessment(*, tutor, exam, bank_item):
    """Copy a tutor-owned bank question into an assessment."""
    _ensure_assessment_owner(tutor=tutor, exam=exam)
    if bank_item.tutor_id != tutor.id:
        raise PermissionError("Question bank item is not owned by this tutor.")
    question = Question.objects.create(
        exam=exam,
        question_type=bank_item.question_type,
        text=bank_item.text,
        explanation=bank_item.explanation,
        media_url=bank_item.media_url,
        marks=bank_item.marks,
        correct_labels=bank_item.correct_labels,
        order=_next_question_order(exam),
        reusable_in_bank=True,
    )
    _create_choices(question=question, choices=bank_item.choices_json)
    _sync_assessment_total_marks(exam)
    return question


def import_questions_from_file(*, tutor, exam, upload):
    """Import assessment questions from CSV/XLSX rows.

    Expected columns: section, question_type, text, marks, correct_labels,
    explanation, media_url, choice_a through choice_f.
    """
    _ensure_assessment_owner(tutor=tutor, exam=exam)
    rows = _read_student_rows(upload)
    result = {"created": [], "errors": []}

    for row_number, row in enumerate(rows, start=2):
        text = (row.get("text") or "").strip()
        question_type = (row.get("question_type") or Question.TYPE_SINGLE_CHOICE).strip()
        if not text:
            result["errors"].append(f"Row {row_number}: question text is required.")
            continue
        if question_type not in dict(Question.QUESTION_TYPE_CHOICES):
            result["errors"].append(f"Row {row_number}: invalid question type.")
            continue

        section = _section_for_import_row(exam=exam, title=(row.get("section") or "").strip())
        question = Question.objects.create(
            exam=exam,
            section=section,
            question_type=question_type,
            text=text,
            explanation=(row.get("explanation") or "").strip(),
            media_url=(row.get("media_url") or "").strip(),
            marks=row.get("marks") or 1,
            correct_labels=(row.get("correct_labels") or "").strip(),
            order=_next_question_order(exam),
        )
        choices = _choices_from_data(row)
        _create_choices(question=question, choices=choices)
        result["created"].append(question)

    _sync_assessment_total_marks(exam)
    return result


def _ensure_assessment_owner(*, tutor, exam):
    """Raise when a tutor attempts to touch another tutor's assessment."""
    if exam.classroom.tutor_id != tutor.id:
        raise PermissionError("Assessment is not owned by this tutor.")


def _choices_from_data(data):
    """Return non-empty A-F choices from form or import data."""
    choices = []
    for label in ["a", "b", "c", "d", "e", "f"]:
        value = data.get(f"choice_{label}") or data.get(label.upper())
        if value:
            choices.append({"label": label.upper(), "text": str(value).strip()})
    return choices


def _create_choices(*, question, choices):
    """Create normalized choices for a question."""
    for choice in choices:
        if choice.get("text"):
            Choice.objects.create(
                question=question,
                label=str(choice["label"]).upper(),
                text=choice["text"],
            )


def _next_question_order(exam):
    """Return the next question order value for an assessment."""
    last_question = exam.questions.order_by("-order", "-id").first()
    return 1 if not last_question else last_question.order + 1


def _section_for_import_row(*, exam, title):
    """Find or create an import section by title."""
    if not title:
        return None
    section, _ = QuestionSection.objects.get_or_create(
        exam=exam,
        title=title,
        defaults={"order": exam.sections.count() + 1},
    )
    return section


def _sync_assessment_total_marks(exam):
    """Keep assessment total marks aligned with question marks."""
    total = sum(question.marks for question in exam.questions.all())
    exam.total_marks = total
    exam.save(update_fields=["total_marks", "updated_at"])


def log_admin_action(*, actor, action, target=None, summary="", metadata=None):
    """Record an admin operation for accountability."""
    return AuditLog.objects.create(
        actor=actor,
        action=action,
        target_model=target.__class__.__name__ if target else "",
        target_id=str(getattr(target, "id", "")) if target else "",
        summary=summary,
        metadata=metadata or {},
    )


def set_user_suspension(*, actor, user, is_suspended):
    """Suspend or reinstate a platform user."""
    if user.is_platform_admin and user.id == actor.id and is_suspended:
        raise ValueError("You cannot suspend your own admin account.")
    user.is_suspended = is_suspended
    user.is_active = not is_suspended
    user.save(update_fields=["is_suspended", "is_active"])
    log_admin_action(
        actor=actor,
        action="user_suspended" if is_suspended else "user_reinstated",
        target=user,
        summary=f"{user.email} {'suspended' if is_suspended else 'reinstated'}.",
    )
    return user


def set_tutor_verification(*, actor, tutor, is_verified):
    """Mark a tutor profile verified or unverified."""
    if tutor.role != User.ROLE_TUTOR:
        raise ValueError("Only tutor accounts can be verified.")
    profile, _ = TutorProfile.objects.get_or_create(user=tutor)
    profile.is_verified = is_verified
    profile.save(update_fields=["is_verified"])
    log_admin_action(
        actor=actor,
        action="tutor_verified" if is_verified else "tutor_unverified",
        target=tutor,
        summary=f"{tutor.email} verification set to {is_verified}.",
    )
    return profile


def save_platform_pricing(*, actor, pricing):
    """Persist platform pricing and audit the change."""
    pricing.save()
    log_admin_action(
        actor=actor,
        action="pricing_updated",
        target=pricing,
        summary=f"Pay-per-student pricing set to {pricing.currency} {pricing.pay_per_student_rate}.",
    )
    return pricing


def save_subscription_plan(*, actor, plan):
    """Persist a subscription plan and audit the change."""
    plan.save()
    log_admin_action(
        actor=actor,
        action="subscription_plan_saved",
        target=plan,
        summary=f"Subscription plan saved: {plan.name}.",
    )
    return plan


def cancel_invoice(*, actor, invoice):
    """Cancel an unpaid invoice from the admin console."""
    if invoice.status == Invoice.STATUS_PAID:
        raise ValueError("Paid invoices cannot be cancelled.")
    invoice.status = Invoice.STATUS_CANCELLED
    invoice.save(update_fields=["status"])
    log_admin_action(actor=actor, action="invoice_cancelled", target=invoice, summary=f"Invoice {invoice.reference} cancelled.")
    return invoice


def admin_mark_invoice_paid(*, actor, invoice):
    """Mark an invoice paid manually and audit the operation."""
    payment = mark_invoice_paid(invoice=invoice, method=Payment.METHOD_MANUAL, provider_reference=f"ADMIN-{invoice.reference}")
    log_admin_action(actor=actor, action="invoice_marked_paid", target=invoice, summary=f"Invoice {invoice.reference} manually marked paid.")
    return payment


def create_support_issue_from_contact(*, actor, contact_message):
    """Convert a public contact message into a trackable support issue."""
    issue, created = SupportIssue.objects.get_or_create(
        contact_message=contact_message,
        defaults={
            "subject": f"Contact from {contact_message.name}",
            "description": contact_message.message,
            "priority": SupportIssue.PRIORITY_NORMAL,
            "status": SupportIssue.STATUS_OPEN,
            "assigned_to": actor if actor.is_platform_admin else None,
        },
    )
    contact_message.is_read = True
    contact_message.save(update_fields=["is_read"])
    if created:
        log_admin_action(actor=actor, action="support_issue_created", target=issue, summary=f"Support issue created from contact message {contact_message.id}.")
    return issue


def save_support_issue(*, actor, issue):
    """Persist support issue changes and close timestamps consistently."""
    if issue.status in [SupportIssue.STATUS_RESOLVED, SupportIssue.STATUS_CLOSED] and not issue.resolved_at:
        issue.resolved_at = timezone.now()
    elif issue.status not in [SupportIssue.STATUS_RESOLVED, SupportIssue.STATUS_CLOSED]:
        issue.resolved_at = None
    issue.save()
    log_admin_action(actor=actor, action="support_issue_updated", target=issue, summary=f"Support issue updated: {issue.subject}.")
    return issue


def save_platform_announcement(*, actor, announcement):
    """Persist an announcement and audit the publication settings."""
    if not announcement.created_by_id:
        announcement.created_by = actor
    announcement.save()
    log_admin_action(actor=actor, action="announcement_saved", target=announcement, summary=f"Announcement saved: {announcement.title}.")
    return announcement


def get_tutor_report_context(*, tutor):
    """Build tutor-level performance, completion, and integrity analytics."""
    classrooms = Classroom.objects.filter(tutor=tutor, is_archived=False).order_by("name")
    exams = Exam.objects.filter(classroom__tutor=tutor).select_related("classroom")
    submissions = Submission.objects.filter(exam__classroom__tutor=tutor)
    completed = submissions.filter(completed=True)
    assigned_students = sum(classroom.students.count() for classroom in classrooms)
    completed_count = completed.count()
    disqualified_count = completed.filter(is_disqualified=True).count()
    proctor_logs = ProctorLog.objects.filter(exam__classroom__tutor=tutor)

    classroom_rows = [
        get_classroom_report_context(tutor=tutor, classroom=classroom)["summary"]
        for classroom in classrooms
    ]
    assessment_rows = [
        _assessment_summary(exam=exam)
        for exam in exams.order_by("-created_at")
    ]

    return {
        "summary": {
            "classroom_count": classrooms.count(),
            "assessment_count": exams.count(),
            "assigned_students": assigned_students,
            "submission_count": submissions.count(),
            "completed_count": completed_count,
            "completion_rate": _percent(completed_count, assigned_students),
            "average_score": _average_score(completed),
            "disqualified_count": disqualified_count,
            "violation_count": proctor_logs.count(),
        },
        "classroom_rows": classroom_rows,
        "assessment_rows": assessment_rows,
        "violation_rows": _violation_summary(proctor_logs),
        "recent_submissions": completed.select_related("student", "student__student_profile", "exam", "exam__classroom").order_by("-submitted_at")[:12],
    }


def get_classroom_report_context(*, tutor, classroom):
    """Build analytics for one tutor-owned classroom."""
    if classroom.tutor_id != tutor.id:
        raise PermissionError("Classroom is not owned by this tutor.")
    exams = Exam.objects.filter(classroom=classroom).order_by("-created_at")
    submissions = Submission.objects.filter(exam__classroom=classroom)
    completed = submissions.filter(completed=True)
    student_count = classroom.students.count()
    completed_count = completed.count()

    return {
        "classroom": classroom,
        "summary": {
            "classroom": classroom,
            "student_count": student_count,
            "assessment_count": exams.count(),
            "submission_count": submissions.count(),
            "completed_count": completed_count,
            "completion_rate": _percent(completed_count, student_count * max(exams.count(), 1)),
            "average_score": _average_score(completed),
            "disqualified_count": completed.filter(is_disqualified=True).count(),
            "violation_count": ProctorLog.objects.filter(exam__classroom=classroom).count(),
        },
        "assessment_rows": [_assessment_summary(exam=exam) for exam in exams],
        "student_rows": _classroom_student_rows(classroom=classroom),
    }


def get_assessment_report_context(*, tutor, exam):
    """Build analytics for one tutor-owned assessment."""
    _ensure_assessment_owner(tutor=tutor, exam=exam)
    submissions = Submission.objects.filter(exam=exam).select_related("student", "student__student_profile")
    completed = submissions.filter(completed=True)
    proctor_logs = ProctorLog.objects.filter(exam=exam)
    return {
        "assessment": exam,
        "summary": _assessment_summary(exam=exam),
        "score_distribution": _score_distribution(completed=completed, exam=exam),
        "question_rows": _question_difficulty_rows(exam=exam),
        "violation_rows": _violation_summary(proctor_logs),
        "submission_rows": completed.order_by("-submitted_at"),
    }


def get_student_report_context(*, student):
    """Build a student's personal result and progress report."""
    overview = get_student_assessment_overview(student=student)
    submissions = Submission.objects.filter(student=student, completed=True).select_related("exam", "exam__classroom")
    completed_count = submissions.count()
    return {
        "assessment_overview": overview,
        "summary": {
            "completed_count": completed_count,
            "pending_count": len(overview["pending"]),
            "active_count": len(overview["active"]),
            "overdue_count": len(overview["overdue"]),
            "average_score": _average_score(submissions),
            "disqualified_count": submissions.filter(is_disqualified=True).count(),
        },
        "submissions": submissions.order_by("-submitted_at"),
    }


def get_admin_report_context():
    """Build platform-level analytics for administrators."""
    paid_invoices = Invoice.objects.filter(status=Invoice.STATUS_PAID)
    pending_invoices = Invoice.objects.filter(status=Invoice.STATUS_PENDING)
    submissions = Submission.objects.all()
    completed = submissions.filter(completed=True)
    return {
        "summary": {
            "tutor_count": User.objects.filter(role=User.ROLE_TUTOR).count(),
            "student_count": User.objects.filter(role=User.ROLE_STUDENT).count(),
            "classroom_count": Classroom.objects.count(),
            "assessment_count": Exam.objects.count(),
            "published_assessment_count": Exam.objects.filter(status=Exam.STATUS_PUBLISHED).count(),
            "submission_count": submissions.count(),
            "completed_count": completed.count(),
            "average_score": _average_score(completed),
            "disqualified_count": completed.filter(is_disqualified=True).count(),
            "violation_count": ProctorLog.objects.count(),
            "paid_revenue": paid_invoices.aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00"),
            "pending_revenue": pending_invoices.aggregate(total=Sum("total_amount"))["total"] or Decimal("0.00"),
            "paid_invoice_count": paid_invoices.count(),
        },
        "top_tutors": _admin_tutor_rows(),
        "recent_submissions": completed.select_related("student", "exam", "exam__classroom").order_by("-submitted_at")[:12],
        "violation_rows": _violation_summary(ProctorLog.objects.all()),
    }


def tutor_assessment_export_rows(*, tutor, exam):
    """Return Excel-compatible CSV rows for an assessment report."""
    context = get_assessment_report_context(tutor=tutor, exam=exam)
    rows = [["Student", "Registration Number", "Email", "Attempt", "Score", "Status", "Submitted", "Violations"]]
    for submission in context["submission_rows"]:
        profile = _student_profile_for_report(student=submission.student)
        rows.append(
            [
                submission.student.fullname,
                profile.registration_number if profile else "",
                submission.student.email,
                submission.attempt_number,
                submission.score,
                "Disqualified" if submission.is_disqualified else "Completed",
                submission.submitted_at,
                submission.proctor_logs.count(),
            ]
        )
    return rows


def _assessment_summary(*, exam):
    submissions = Submission.objects.filter(exam=exam)
    completed = submissions.filter(completed=True)
    completed_count = completed.count()
    assigned_count = exam.classroom.students.count()
    passed_count = completed.filter(is_disqualified=False, score__gte=float(exam.pass_mark)).count()
    score_bounds = completed.aggregate(highest=Max("score"), lowest=Min("score"))
    return {
        "assessment": exam,
        "assigned_count": assigned_count,
        "submission_count": submissions.count(),
        "completed_count": completed_count,
        "completion_rate": _percent(completed_count, assigned_count),
        "average_score": _average_score(completed),
        "highest_score": score_bounds["highest"],
        "lowest_score": score_bounds["lowest"],
        "passed_count": passed_count,
        "pass_rate": _percent(passed_count, completed_count),
        "disqualified_count": completed.filter(is_disqualified=True).count(),
        "violation_count": ProctorLog.objects.filter(exam=exam).count(),
    }


def _classroom_student_rows(*, classroom):
    rows = []
    for student in classroom.students.select_related("student_profile").order_by("fullname", "email"):
        profile = _student_profile_for_report(student=student)
        submissions = Submission.objects.filter(student=student, exam__classroom=classroom, completed=True)
        rows.append(
            {
                "student": student,
                "profile": profile,
                "completed_count": submissions.count(),
                "average_score": _average_score(submissions),
                "disqualified_count": submissions.filter(is_disqualified=True).count(),
                "violation_count": ProctorLog.objects.filter(student=student, exam__classroom=classroom).count(),
            }
        )
    return rows


def _student_profile_for_report(*, student):
    """Return a student's profile for report identity fields when available."""
    try:
        return student.student_profile
    except StudentProfile.DoesNotExist:
        return None


def _question_difficulty_rows(*, exam):
    rows = []
    completed_submissions = Submission.objects.filter(exam=exam, completed=True)
    for question in exam.questions.order_by("section__order", "order", "id"):
        answers = StudentAnswer.objects.filter(submission__in=completed_submissions, question=question)
        answer_count = answers.count()
        correct_count = answers.filter(is_correct=True).count()
        rows.append(
            {
                "question": question,
                "answer_count": answer_count,
                "correct_count": correct_count,
                "correct_rate": _percent(correct_count, answer_count),
                "average_marks": answers.aggregate(value=Avg("awarded_marks"))["value"] or Decimal("0.00"),
            }
        )
    return rows


def _score_distribution(*, completed, exam):
    buckets = [
        {"label": "0-39", "minimum": 0, "maximum": 39, "count": 0},
        {"label": "40-49", "minimum": 40, "maximum": 49, "count": 0},
        {"label": "50-59", "minimum": 50, "maximum": 59, "count": 0},
        {"label": "60-69", "minimum": 60, "maximum": 69, "count": 0},
        {"label": "70-100", "minimum": 70, "maximum": 100, "count": 0},
    ]
    total_marks = float(exam.total_marks or 0)
    for submission in completed:
        percent_score = int(0 if total_marks <= 0 else min(100, max(0, (submission.score / total_marks) * 100)))
        for bucket in buckets:
            if bucket["minimum"] <= percent_score <= bucket["maximum"]:
                bucket["count"] += 1
                break
    return buckets


def _violation_summary(logs):
    labels = dict(ProctorLog.VIOLATION_TYPES)
    return [
        {
            "violation_type": row["violation_type"],
            "label": labels.get(row["violation_type"], row["violation_type"]),
            "count": row["count"],
            "disqualification_count": row["disqualification_count"],
        }
        for row in logs.values("violation_type")
        .annotate(count=Count("id"), disqualification_count=Count("id", filter=Q(triggered_disqualification=True)))
        .order_by("-count")
    ]


def _admin_tutor_rows():
    rows = []
    for tutor in User.objects.filter(role=User.ROLE_TUTOR).order_by("fullname", "email")[:20]:
        exams = Exam.objects.filter(classroom__tutor=tutor)
        completed = Submission.objects.filter(exam__classroom__tutor=tutor, completed=True)
        rows.append(
            {
                "tutor": tutor,
                "classroom_count": Classroom.objects.filter(tutor=tutor).count(),
                "assessment_count": exams.count(),
                "student_count": StudentProfile.objects.filter(tutor=tutor).count(),
                "completed_count": completed.count(),
                "average_score": _average_score(completed),
                "violation_count": ProctorLog.objects.filter(exam__classroom__tutor=tutor).count(),
            }
        )
    return rows


def _average_score(queryset):
    return queryset.aggregate(value=Avg("score"))["value"] or 0


def _percent(numerator, denominator):
    if not denominator:
        return 0
    return round((numerator / denominator) * 100, 1)
